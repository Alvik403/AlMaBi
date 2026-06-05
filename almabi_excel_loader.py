from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from almabi_mock_data import get_almabi_dashboard_data
from almabi_template_data import get_almabi_template_dashboard_data


_MONTH_NAMES = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def _parse_amount(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or value == "":
        return None
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _find_header_row(sheet) -> tuple[int, dict[str, int]] | None:
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=30, values_only=True),
        start=1,
    ):
        headers = [_normalize_header(value) for value in row]
        column_map: dict[str, int] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            column_map[header] = index
        if {"документ", "сумма"} <= set(column_map):
            return row_index, column_map
        if {"document", "sum buh"} <= set(column_map):
            return row_index, column_map
        if "выручка" in column_map or "revenue" in column_map:
            return row_index, column_map
    return None


def _aggregate_bdr_rows(sheet, header_row: int, column_map: dict[str, int]) -> dict[str, float]:
    date_idx = column_map.get("дата", column_map.get("date"))
    amount_idx = column_map.get("сумма", column_map.get("sum buh"))
    if amount_idx is None:
        return {}

    totals = {month: 0.0 for month in _MONTH_NAMES.values()}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        amount = _parse_amount(row[amount_idx] if amount_idx < len(row) else None)
        if not amount:
            continue
        parsed_date = None
        if date_idx is not None and date_idx < len(row):
            parsed_date = _parse_date(row[date_idx])
        if parsed_date is None:
            continue
        month_name = _MONTH_NAMES.get(parsed_date.month)
        if month_name:
            totals[month_name] += amount
    return totals


def _aggregate_realization_rows(sheet, header_row: int, column_map: dict[str, int]) -> dict[str, float]:
    date_idx = column_map.get("регистратор.дата", column_map.get("date"))
    amount_idx = column_map.get("выручка", column_map.get("revenue"))
    if amount_idx is None:
        return {}

    totals = {month: 0.0 for month in _MONTH_NAMES.values()}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        amount = _parse_amount(row[amount_idx] if amount_idx < len(row) else None)
        if not amount:
            continue
        parsed_date = None
        if date_idx is not None and date_idx < len(row):
            parsed_date = _parse_date(row[date_idx])
        if parsed_date is None:
            continue
        month_name = _MONTH_NAMES.get(parsed_date.month)
        if month_name:
            totals[month_name] += amount
    return totals


def load_almabi_dashboard_from_excel(path: Path, *, original_name: str) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        header_info = None
        monthly_totals: dict[str, float] = {}
        for sheet in workbook.worksheets:
            header_info = _find_header_row(sheet)
            if not header_info:
                continue
            header_row, column_map = header_info
            if "выручка" in column_map or "revenue" in column_map:
                monthly_totals = _aggregate_realization_rows(sheet, header_row, column_map)
            else:
                monthly_totals = _aggregate_bdr_rows(sheet, header_row, column_map)
            if any(monthly_totals.values()):
                break
    finally:
        workbook.close()

    if not any(monthly_totals.values()):
        data = get_almabi_template_dashboard_data()
        data["meta"] = {
            **data.get("meta", {}),
            "source": "upload",
            "upload_file_name": original_name,
            "parsed": False,
            "message": (
                "Файл принят, но в нём нет строк с суммами по месяцам. "
                "Показаны тестовые данные шаблона БДР."
            ),
        }
        return data

    data = get_almabi_dashboard_data()
    total_revenue = sum(monthly_totals.values())
    scale = total_revenue / max(data["cost_structure_total"], 1)
    data["meta"] = {
        "source": "upload",
        "upload_file_name": original_name,
        "parsed": True,
        "message": "Данные агрегированы из загруженного файла (выручка по месяцам).",
    }
    data["revenue_by_month"] = [
        {"month": month[:4] + ".", "value": round(monthly_totals.get(full_month, 0))}
        for full_month, month in zip(
            _MONTH_NAMES.values(),
            ["Янв.", "Фев.", "Мар.", "Апр.", "Май", "Июн.", "Июл.", "Авг.", "Сен.", "Окт.", "Ноя.", "Дек."],
            strict=True,
        )
    ]
    data["cost_by_month"] = [
        {"month": item["month"], "value": round(item["value"] * 0.82 * scale)}
        for item in data["revenue_by_month"]
    ]
    data["cost_structure_total"] = sum(item["value"] for item in data["cost_by_month"])
    return data
