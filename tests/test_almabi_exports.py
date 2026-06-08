from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from almabi_dashboard_builder import load_almabi_dashboard_from_exports
from almabi_file_validation import validate_almabi_export


def _pad_rows(sheet, count: int) -> None:
    for _ in range(count):
        sheet.append([None])


def create_buh_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист_1"
    _pad_rows(sheet, 8)
    headers = [
        "Документ",
        "Счет Дт",
        "Вид субконто2 Дт",
        "Субконто2 Дт",
        "Счет Кт",
        "Вид субконто1 Кт",
        "Субконто1 Кт",
        "Дата",
        "Договор",
        "Сумма",
        "Сумма НУ Дт",
        "Сумма НУ Кт",
    ]
    sheet.append(headers)
    realization_document = "Реализация товаров и услуг 00АМ-000017 от 31.01.2026 21:00:00"
    sheet.append(
        [
            realization_document,
            "90.02.1",
            "Варианты налогообложения прибыли",
            "Общие условия налогообложения",
            "43",
            "Номенклатура",
            "Лицензия ПО",
            "15.01.2026",
            "Д-001",
            400_000,
            400_000,
            0,
        ]
    )
    sheet.append(
        [
            realization_document,
            "62.01",
            "Варианты налогообложения прибыли",
            "Доходы по льготируемым видам деятельности",
            "90.01.3",
            "",
            "",
            "15.01.2026",
            "Д-001",
            1_000_000,
            0,
            1_000_000,
        ]
    )
    sheet.append(
        [
            realization_document,
            "62.01",
            "Контрагенты",
            "ООО Тест Клиент",
            "51",
            "",
            "",
            "15.01.2026",
            "Д-001",
            1,
            0,
            0,
        ]
    )
    sheet.append(
        [
            realization_document,
            "90.07.1",
            "Статьи затрат",
            "Реклама",
            "44",
            "",
            "",
            "20.02.2026",
            "",
            50_000,
            50_000,
            0,
        ]
    )
    sheet.append(["Итого"])
    workbook.save(path)


def create_realization_workbook(path: Path, *, document: str = "Реализация товаров и услуг 00АМ-000017 от 31.01.2026 21:00:00") -> None:
    workbook = Workbook()
    sheet = workbook.active
    _pad_rows(sheet, 7)
    headers = [
        "Заказ клиента / Реализация",
        "Номенклатура",
        "Проект",
        "Группа проектов",
        "Направление",
        "Выручка",
        "Валовая прибыль",
    ]
    sheet.append(headers)
    sheet.append(
        [
            document,
            "Лицензия ПО",
            "Обслуживание Долго",
            "Обслуживание",
            "Услуги",
            1_000_000,
            600_000,
        ]
    )
    sheet.append(["Итого"])
    workbook.save(path)


def create_cost_workbook(path: Path, *, document: str = "Реализация товаров и услуг 00АМ-000017 от 31.01.2026 21:00:00") -> None:
    workbook = Workbook()
    sheet = workbook.active
    _pad_rows(sheet, 5)
    headers = [
        "Продукция",
        "Счет",
        "Статья калькуляции",
        "Документ отгрузки",
        "Количество продаж",
        "Себестоимость (бухг. учет)",
    ]
    sheet.append(headers)
    sheet.append(
        [
            "Лицензия ПО",
            "20",
            "Материальные затраты",
            document,
            1,
            400_000,
        ]
    )
    _pad_rows(sheet, 38)
    workbook.save(path)


def create_amort_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    _pad_rows(sheet, 8)
    headers = [
        "Статья расходов",
        "Организация",
        "Подразделение",
        "Статья калькуляции",
        "Регистратор (рег.) приход",
        "Дата записи",
        "Направление деятельности",
        "Стоимость",
    ]
    sheet.append(headers)
    sheet.append(["Амортизация ОС", "ООО Тест", "Цех 1", "Амортизация", "Док 1", "31.01.2026", "Услуги", 25_000])
    sheet.append(["Итого"])
    workbook.save(path)


def test_validate_export_types(tmp_path: Path):
    paths = {
        "buh": tmp_path / "buh.xlsx",
        "realization": tmp_path / "realization.xlsx",
        "cost": tmp_path / "cost.xlsx",
        "amortization": tmp_path / "amort.xlsx",
    }
    create_buh_workbook(paths["buh"])
    create_realization_workbook(paths["realization"])
    create_cost_workbook(paths["cost"])
    create_amort_workbook(paths["amortization"])

    assert validate_almabi_export(paths["buh"]).export_type == "buh"
    assert validate_almabi_export(paths["realization"]).export_type == "realization"
    assert validate_almabi_export(paths["cost"]).export_type == "cost"
    assert validate_almabi_export(paths["amortization"]).export_type == "amortization"


def test_realization_projects_columns_are_parsed(tmp_path: Path):
    path = tmp_path / "realization.xlsx"
    create_realization_workbook(path)
    from almabi_export_parsers import parse_realization

    rows = parse_realization(path)
    assert rows[0].direction == "Услуги"
    assert rows[0].project_group == "Обслуживание"
    assert rows[0].project == "Обслуживание Долго"
    assert rows[0].month == "Январь"


def test_project_index_matches_documents_by_number(tmp_path: Path):
    from almabi_pipeline import build_facts
    from almabi_export_parsers import parse_exports

    realization_path = tmp_path / "realization.xlsx"
    cost_path = tmp_path / "cost.xlsx"
    create_realization_workbook(
        realization_path,
        document="Реализация товаров и услуг 00АМ-000017 от 31.01.2026 21:00:00",
    )
    create_cost_workbook(cost_path)

    exports = parse_exports({"realization": realization_path, "cost": cost_path})
    result = build_facts(exports)
    revenue = [fact for fact in result.facts if fact.kpi_l1 == "Выручка"]
    assert revenue
    assert revenue[0].direction == "Услуги"
    assert revenue[0].project_group == "Обслуживание"


def test_dashboard_builder_from_exports(tmp_path: Path):
    paths = {
        "buh": tmp_path / "buh.xlsx",
        "realization": tmp_path / "realization.xlsx",
        "cost": tmp_path / "cost.xlsx",
    }
    create_buh_workbook(paths["buh"])
    create_realization_workbook(paths["realization"])
    create_cost_workbook(paths["cost"])

    dashboard = load_almabi_dashboard_from_exports(
        paths,
        upload_names={key: path.name for key, path in paths.items()},
    )

    assert dashboard["meta"]["parsed"] is True
    assert len(dashboard["summary_rows"]) == 10
    revenue = next(row for row in dashboard["summary_rows"] if row["name"] == "Выручка")
    assert revenue["total_fact"] == 1_000_000
    cost = next(row for row in dashboard["summary_rows"] if row["name"] == "Себестоимость")
    assert cost["total_fact"] == -400_000
    assert "Услуги" in {child["name"] for child in revenue["children"]}
    assert dashboard["contractor_details"]
    assert dashboard["contractor_cards"]
    assert any(item["contractor"] == "ООО Тест Клиент" for item in dashboard["contractor_details"])
    privileged = next(card for card in dashboard["contractor_cards"] if "льгот" in card["title"].casefold() and "нельгот" not in card["title"].casefold())
    assert privileged["total"] == 1_000_000
    assert privileged["rows"][0]["name"] == "ООО Тест Клиент"


def test_amortization_is_not_detected_as_cost(tmp_path: Path):
    path = tmp_path / "amort.xlsx"
    create_amort_workbook(path)
    result = validate_almabi_export(path)
    assert result.export_type == "amortization"


def test_upload_bundle_accepts_misplaced_export_file(app_client, tmp_path: Path):
    files = {
        "buh_file": ("buh.xlsx", _workbook_bytes(create_buh_workbook, tmp_path / "buh.xlsx")),
        "realization_file": ("realization.xlsx", _workbook_bytes(create_realization_workbook, tmp_path / "realization.xlsx")),
        "amortization_file": ("cost.xlsx", _workbook_bytes(create_cost_workbook, tmp_path / "cost.xlsx")),
    }

    response = app_client.post("/api/almabi/files/upload-set", files=files)
    assert response.status_code == 201
    payload = response.json()
    assert payload["saved_exports"]["cost"]["original"] == "cost.xlsx"
    assert payload["saved_exports"]["cost"]["selected_slot"] == "amortization"
    assert any("определён как" in warning for warning in payload.get("warnings", []))


def test_upload_bundle_builds_dashboard(app_client, tmp_path: Path):
    files = {
        "buh_file": ("buh.xlsx", _workbook_bytes(create_buh_workbook, tmp_path / "buh.xlsx")),
        "realization_file": ("realization.xlsx", _workbook_bytes(create_realization_workbook, tmp_path / "realization.xlsx")),
        "cost_file": ("cost.xlsx", _workbook_bytes(create_cost_workbook, tmp_path / "cost.xlsx")),
    }

    response = app_client.post("/api/almabi/files/upload-set", files=files)
    assert response.status_code == 201
    payload = response.json()
    assert payload["is_complete"] is True
    assert set(payload["upload_files"]) == {"buh", "realization", "cost"}

    dashboard = app_client.get("/dashboard/almabi")
    assert dashboard.status_code == 200
    assert "Выручка" in dashboard.text
    assert "Услуги" in dashboard.text
    assert "ООО Тест Клиент" in dashboard.text


def _workbook_bytes(factory, path: Path) -> bytes:
    factory(path)
    return path.read_bytes()
