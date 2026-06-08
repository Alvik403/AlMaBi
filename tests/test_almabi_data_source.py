from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


def _create_bdr_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист_1"
    headers = [
        "Документ",
        "Счет Дт",
        "Вид субконто1 Дт",
        "Субконто1 Дт",
        "Счет Кт",
        "Дата",
        "Подразделение Дт",
        "Договор",
        "Сумма",
        "Сумма НУ Дт",
        "Сумма НУ Кт",
    ]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=9, column=index, value=header)
    sheet.cell(row=10, column=1, value="Проводка 1")
    sheet.cell(row=10, column=6, value="15.01.2026")
    sheet.cell(row=10, column=9, value=1_250_000)
    sheet.cell(row=11, column=1, value="Проводка 2")
    sheet.cell(row=11, column=6, value="20.02.2026")
    sheet.cell(row=11, column=9, value=980_000)
    workbook.save(path)


def test_almabi_data_source_defaults_to_mock(app_client):
    response = app_client.get("/api/almabi/data-source")

    assert response.status_code == 200
    payload = response.json()
    assert payload["source"] == "mock"
    assert payload["title"] == "Демо-данные"


def test_almabi_can_switch_to_template(app_client):
    response = app_client.post("/api/almabi/data-source/template")

    assert response.status_code == 200
    assert response.json()["source"] == "template"

    dashboard = app_client.get("/dashboard/almabi")
    assert dashboard.status_code == 200
    assert "Шаблон БДР" in dashboard.text
    assert "П-101" in dashboard.text


def test_almabi_upload_accepts_bdr_file(app_client, tmp_path: Path):
    upload_path = tmp_path / "my_bdr.xlsx"
    _create_bdr_workbook(upload_path)

    with upload_path.open("rb") as workbook:
        response = app_client.post(
            "/api/almabi/files/upload",
            files={
                "file": (
                    "my_bdr.xlsx",
                    workbook,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 201
    payload = response.json()
    assert payload["source"] == "upload"
    assert payload["saved_exports"]["buh"]["original"] == "my_bdr.xlsx"
    assert payload["saved_exports"]["buh"]["validation"]["export_type"] == "buh"

    dashboard = app_client.get("/dashboard/almabi")
    assert dashboard.status_code == 200
    assert "my_bdr.xlsx" in dashboard.text
