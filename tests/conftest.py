from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


def create_sample_workbook(path: Path, *, class_name: str = "экспорт/сибур", service_name: str = "Комплексная/услуга 2") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Калькуляция"

    top_headers = [
        "№",
        "Класс",
        "Услуга",
        "Комментарий",
        "Прямые расходы",
        "Прямые расходы",
        "Прямые расходы",
        "Прямые расходы",
        "Прямые расходы",
        "Прямые расходы",
        "Прямые расходы",
        "Прямые расходы",
        "Прямые расходы",
        "Прямые расходы",
        "Прямые расходы",
        "Прямые расходы",
        "Косвенные расходы",
        "Косвенные расходы",
        "Косвенные расходы",
        "Косвенные расходы",
        "Косвенные расходы",
        "Косвенные расходы",
        "Неэффективность",
        "Неэффективность",
        "Неэффективность",
    ]
    sub_headers = [
        "",
        "",
        "",
        "",
        "ИТОГО",
        "ФОТ",
        "Материалы",
        "Процент %",
        "Прочее",
        "Деталь 5",
        "Деталь 6",
        "Деталь 7",
        "Деталь 8",
        "Деталь 9",
        "Деталь 10",
        "Деталь 11",
        "ИТОГО",
        "Аренда",
        "Связь",
        "Охрана",
        "Прочее",
        "Деталь косвенная",
        "Аренда",
        "РЖД",
        "ПРТ",
    ]
    values = [
        1,
        class_name,
        service_name,
        "",
        1000,
        600,
        400,
        10,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        500,
        200,
        100,
        50,
        50,
        100,
        100,
        50,
        25,
    ]

    for col_idx, value in enumerate(top_headers, start=1):
        sheet.cell(row=3, column=col_idx, value=value)
    for col_idx, value in enumerate(sub_headers, start=1):
        sheet.cell(row=4, column=col_idx, value=value)
    for col_idx, value in enumerate(values, start=1):
        sheet.cell(row=5, column=col_idx, value=value)

    workbook.save(path)


@pytest.fixture()
def sample_excel_path(tmp_path: Path) -> Path:
    path = tmp_path / "data.xlsx"
    create_sample_workbook(path)
    return path


@pytest.fixture()
def app_client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> TestClient:
    data_dir = tmp_path / "data"
    create_sample_workbook(data_dir / "data.xlsx")

    monkeypatch.setenv("DEBUG", "true")
    monkeypatch.setenv("DATA_DIR", str(data_dir))
    monkeypatch.setenv("UPLOADS_DIR", str(tmp_path / "uploads"))
    monkeypatch.setenv("DATABASE_PATH", str(tmp_path / "runtime" / "test.sqlite3"))
    monkeypatch.setenv("LOGS_DIR", str(tmp_path / "logs"))
    monkeypatch.setenv("SESSION_SECRET", "test-secret")

    for module_name in [
        "app",
        "settings",
        "file_registry",
        "dashboard_cache",
        "logging_config",
    ]:
        sys.modules.pop(module_name, None)

    spec = importlib.util.spec_from_file_location("app", PROJECT_ROOT / "app.py")
    assert spec and spec.loader
    app_module = importlib.util.module_from_spec(spec)
    sys.modules["app"] = app_module
    spec.loader.exec_module(app_module)
    with TestClient(app_module.app) as client:
        yield client
