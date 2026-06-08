from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook

EMPTY_ANALYTICS = {"", "без направления", "без группы", "без проекта", "без договора", "без номенклатуры"}

MONTH_NAMES = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_amount(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or value == "":
        return None
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def month_name(value: date | None) -> str | None:
    if value is None:
        return None
    return MONTH_NAMES.get(value.month)


def build_column_map(headers: Iterable[object]) -> dict[str, int]:
    column_map: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        if normalized:
            column_map[normalized] = index
    return column_map


def header_matches(header: str, candidate: str) -> bool:
    header_value = normalize_header(header)
    candidate_value = candidate.casefold()
    if not header_value or not candidate_value:
        return False
    if header_value == candidate_value:
        return True
    if header_value.endswith(f".{candidate_value}"):
        return True
    if f"[{candidate_value}]" in header_value:
        return True
    for part in re.split(r"\s*/\s*", header_value):
        if part == candidate_value or part.endswith(f".{candidate_value}"):
            return True
    tokens = re.split(r"[.\[\]/]+", header_value)
    return candidate_value in tokens


def parse_date_from_text(value: object) -> date | None:
    text = normalize_text(value)
    if not text:
        return None
    direct = parse_date(text)
    if direct:
        return direct
    for pattern in (
        r"от\s+(\d{2}\.\d{2}\.\d{4})",
        r"(\d{2}\.\d{2}\.\d{4})",
        r"(\d{4}-\d{2}-\d{2})",
    ):
        match = re.search(pattern, text)
        if match:
            parsed = parse_date(match.group(1))
            if parsed:
                return parsed
    return None


def resolve_column_map(headers: list[str], aliases: dict[str, tuple[str, ...]]) -> dict[str, int]:
    normalized_headers = [normalize_header(header) for header in headers]
    resolved: dict[str, int] = {}
    for target, options in aliases.items():
        for index, header in enumerate(normalized_headers):
            if any(header_matches(header, option) for option in options):
                resolved[target] = index
                break
    return resolved


def cell_value(row: tuple[object, ...], column_map: dict[str, int], *keys: str) -> object:
    for key in keys:
        index = column_map.get(key)
        if index is None or index >= len(row):
            continue
        return row[index]
    return None


def analytics_value(value: object, *, default: str) -> str:
    text = normalize_text(value)
    if not text or text.casefold() in EMPTY_ANALYTICS:
        return default
    return text


def document_match_keys(value: str) -> set[str]:
    text = normalize_text(value)
    if not text:
        return set()

    keys = {text.casefold()}
    compact = re.sub(r"\s+", " ", text.casefold())
    keys.add(compact)

    for pattern in (
        r"\d{2}[a-zа-я]{2}-\d+",
        r"[a-zа-я]{2,}-\d+/\d+",
        r"договор\s*№?\s*([a-zа-я0-9\-/]+)",
        r"реализация[^0-9]*(\d{2}[a-zа-я]{2}-\d+)",
    ):
        match = re.search(pattern, compact, flags=re.IGNORECASE)
        if match:
            keys.add(match.group(1 if match.lastindex else 0).casefold())
    return keys


def read_workbook_rows(path: Path) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        return list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def read_sheet_rows(
    path: Path,
    *,
    skip_rows: int = 0,
    remove_last: int = 0,
    header_matcher: Callable[[list[str]], bool] | None = None,
    scan_rows: int = 40,
) -> tuple[list[str], list[tuple[object, ...]]]:
    raw_rows = read_workbook_rows(path)
    if remove_last and len(raw_rows) > remove_last:
        raw_rows = raw_rows[:-remove_last]

    header_index = 0
    if header_matcher:
        for index, row in enumerate(raw_rows[:scan_rows]):
            headers = [normalize_text(value) for value in row]
            if header_matcher([normalize_header(value) for value in headers if value is not None]):
                header_index = index
                break
    elif skip_rows:
        header_index = skip_rows

    raw_rows = raw_rows[header_index:]
    if not raw_rows:
        return [], []

    headers = [normalize_text(value) for value in raw_rows[0]]
    data_rows = raw_rows[1:]
    return headers, data_rows


def find_subconto_value(
    column_map: dict[str, int],
    row: tuple[object, ...],
    *,
    side: str,
    kind_marker: str,
) -> str:
    marker = kind_marker.casefold()
    for level in (1, 2, 3):
        kind = normalize_text(cell_value(row, column_map, f"вид субконто{level} {side}"))
        if marker in kind:
            return normalize_text(cell_value(row, column_map, f"субконто{level} {side}"))
    return ""


def tax_bucket(tax_type: str) -> str:
    normalized = tax_type.casefold()
    if not tax_type or "общие условия" in normalized:
        return "Нельготные проекты"
    if any(marker in normalized for marker in ("льгот", "льготиру", "пониженн", "ставка 0")):
        return "Льготные проекты"
    return "Нельготные проекты"
