from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from almabi_excel_utils import normalize_header


EXPORT_TYPES = ("buh", "realization", "cost", "amortization")
REQUIRED_EXPORT_TYPES = ("buh", "realization", "cost")

EXPORT_LABELS = {
    "buh": "Бух.регистр",
    "realization": "Реализация",
    "cost": "Себестоимость",
    "amortization": "Амортизация",
}


@dataclass(frozen=True)
class AlmabiValidationResult:
    export_type: str
    sheet_name: str
    header_row: int

    def to_dict(self) -> dict:
        return {
            "export_type": self.export_type,
            "format": self.export_type,
            "sheet_name": self.sheet_name,
            "header_row": self.header_row,
            "label": EXPORT_LABELS[self.export_type],
        }


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_texts(row: tuple[object, ...]) -> list[str]:
    return [_normalize_cell(value) for value in row if _normalize_cell(value)]


def _joined_headers(cells: list[str]) -> str:
    return " | ".join(normalize_header(cell) for cell in cells)


def guess_export_type_from_filename(filename: str) -> str | None:
    normalized = Path(filename).name.casefold()
    if any(token in normalized for token in ("аморт", "amort")):
        return "amortization"
    if any(token in normalized for token in ("себест", "cost", "себестоим")):
        return "cost"
    if any(token in normalized for token in ("реализ", "realiz", "выручк")):
        return "realization"
    if any(token in normalized for token in ("бух", "buh", "регистр")):
        return "buh"
    return None


def _detect_export_type(cells: list[str]) -> str | None:
    joined = _joined_headers(cells)
    if ("документ отгрузки" in joined or ("продукция" in joined and "себестоимость" in joined)) and (
        "себестоимость (бухг" in joined or "себестоимость (регл" in joined or "себестоимость полная" in joined
    ):
        return "cost"
    if "статья расходов" in joined and "направление деятельности" in joined:
        return "amortization"
    if "статья расходов" in joined and "регистратор" in joined and "документ отгрузки" not in joined:
        return "amortization"
    if "выручка" in joined and ("номенклатура" in joined or "группа проектов" in joined or "заказ клиента" in joined):
        return "realization"
    if all(marker in joined for marker in ("документ", "сумма", "счет дт")):
        return "buh"
    return None


def validate_almabi_export(path: Path, *, expected_type: str | None = None) -> AlmabiValidationResult:
    if path.suffix.casefold() != ".xlsx":
        raise ValueError("Поддерживаются только файлы .xlsx")
    if not path.exists():
        raise ValueError(f"Файл не найден: {path}")
    if path.stat().st_size == 0:
        raise ValueError("Файл пустой")
    if expected_type and expected_type not in EXPORT_TYPES:
        raise ValueError(f"Неизвестный тип выгрузки: {expected_type}")

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Не удалось открыть Excel: {exc}") from exc

    try:
        for sheet in workbook.worksheets:
            for row_index, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=40, values_only=True),
                start=1,
            ):
                cells = _row_texts(row)
                if not cells:
                    continue
                detected = _detect_export_type(cells)
                if not detected:
                    continue
                if expected_type and detected != expected_type:
                    raise ValueError(
                        f"Ожидалась выгрузка «{EXPORT_LABELS[expected_type]}», "
                        f"но файл похож на «{EXPORT_LABELS[detected]}»."
                    )
                return AlmabiValidationResult(
                    export_type=detected,
                    sheet_name=sheet.title,
                    header_row=row_index,
                )
    finally:
        workbook.close()

    if expected_type:
        raise ValueError(
            f"Файл не похож на выгрузку «{EXPORT_LABELS[expected_type]}». "
            "Проверьте структуру колонок."
        )
    raise ValueError(
        "Файл не похож на одну из выгрузок БДР. "
        "Ожидаются колонки бухрегистра, реализации, себестоимости или амортизации."
    )


def validate_almabi_excel(path: Path) -> AlmabiValidationResult:
    return validate_almabi_export(path)
