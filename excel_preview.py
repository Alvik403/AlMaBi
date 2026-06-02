from __future__ import annotations

import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet

# Размер «окна» на экране (не весь лист — навигация стрелками / ссылками).
DEFAULT_VIEWPORT_ROWS = 45
DEFAULT_VIEWPORT_COLS = 28


def _cell_text(cell: Cell | None) -> str:
    if cell is None:
        return ""
    value = cell.value
    if value is None:
        return ""
    return str(value)


def _column_width_ch(ws: Worksheet, col_idx: int) -> str | None:
    """Грубая оценка ширины колонки для CSS (ch), если задана в openpyxl."""
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    if dim is None or dim.width is None:
        return None
    try:
        w = float(dim.width)
    except (TypeError, ValueError):
        return None
    ch = max(4, min(48, int(math.ceil(w))))
    return f"{ch}ch"


@dataclass(frozen=True)
class PreviewCell:
    """Одна ячейка HTML-таблицы (после учёта merge)."""

    text: str
    sheet_row: int
    sheet_col: int
    rowspan: int = 1
    colspan: int = 1
    skip: bool = False
    col_width: str | None = None


def _resolve_merge_anchor(ws: Worksheet, row: int, column: int) -> tuple[int, int]:
    """Левый верх объединения, если ячейка входит в merge; иначе (row, column)."""
    merged = getattr(ws, "merged_cells", None)
    if merged is None:
        return row, column
    for mrange in merged.ranges:
        min_row, min_col, max_row, max_col = mrange.bounds
        if min_row <= row <= max_row and min_col <= column <= max_col:
            return min_row, min_col
    return row, column


def get_cell_detail(path: Path, sheet_index: int, row: int, column: int) -> dict[str, Any]:
    """
    Содержимое одной ячейки: как в книге и кэш пересчёта Excel (если сохранён).
    Открывает файл дважды (data_only False/True): полноценный пересчёт формул на сервере не выполняется.
    """
    if row < 1 or column < 1:
        raise ValueError("Строка и столбец должны быть ≥ 1")

    wb_formula = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    wb_values = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if sheet_index < 0 or sheet_index >= len(wb_formula.worksheets):
            raise ValueError(f"Лист {sheet_index} не найден")
        ws_f = wb_formula.worksheets[sheet_index]
        ws_v = wb_values.worksheets[sheet_index]
        ar, ac = _resolve_merge_anchor(ws_f, row, column)
        cf = ws_f.cell(row=ar, column=ac)
        cv = ws_v.cell(row=ar, column=ac)
        raw = cf.value
        evaluated = cv.value
        is_formula = cf.data_type == "f" or (isinstance(raw, str) and raw.startswith("="))
        address = f"{get_column_letter(ac)}{ar}"
        contents = None if raw is None else str(raw)
        evaluated_repr: str | None
        if evaluated is None:
            evaluated_repr = None
        elif isinstance(evaluated, float) and math.isnan(evaluated):
            evaluated_repr = "NaN"
        else:
            evaluated_repr = str(evaluated)

        merge_note: str | None = None
        if (ar, ac) != (row, column):
            merge_note = f"Ячейка входит в объединение; значение показано для {address}"

        return {
            "address": address,
            "row": ar,
            "column": ac,
            "clicked_row": row,
            "clicked_column": column,
            "contents": contents,
            "is_formula": is_formula,
            "evaluated_value": evaluated_repr,
            "cached_value_missing": bool(is_formula and evaluated is None),
            "merge_note": merge_note,
        }
    finally:
        wb_formula.close()
        wb_values.close()


_CELL_REF_RE = re.compile(r"^[A-Za-z]+\d+$")


def parse_goto_cell(goto: str | None) -> tuple[int | None, int | None]:
    """
    Простой разбор A1 или a1; буквы столбец, остальное — строка.
    При ошибке (None, None).
    """
    if not goto:
        return None, None
    token = goto.strip().upper()
    if not _CELL_REF_RE.match(token):
        return None, None
    try:
        col_letters, row_num = coordinate_from_string(token)
    except Exception:
        return None, None
    try:
        col_idx = column_index_from_string(col_letters)
    except Exception:
        return None, None
    if row_num < 1 or col_idx < 1:
        return None, None
    return row_num, col_idx


def _apply_merges_viewport(
    ws: Worksheet,
    vr1: int,
    vr2: int,
    vc1: int,
    vc2: int,
) -> tuple[dict[tuple[int, int], tuple[int, int, int, int]], set[tuple[int, int]]]:
    spans: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    skip: set[tuple[int, int]] = set()

    for mrange in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = mrange.bounds
        vis_min_r = max(min_row, vr1)
        vis_min_c = max(min_col, vc1)
        vis_max_r = min(max_row, vr2)
        vis_max_c = min(max_col, vc2)
        if vis_min_r > vis_max_r or vis_min_c > vis_max_c:
            continue

        if min_row > vr2 or min_col > vc2:
            continue

        rowspan = vis_max_r - vis_min_r + 1
        colspan = vis_max_c - vis_min_c + 1
        top_left = (vis_min_r, vis_min_c)
        spans[top_left] = (rowspan, colspan, min_row, min_col)

        for r in range(vis_min_r, vis_max_r + 1):
            for c in range(vis_min_c, vis_max_c + 1):
                if (r, c) != top_left:
                    skip.add((r, c))

    return spans, skip


def read_sheet_bounds(path: Path, sheet_index: int) -> tuple[list[str], int, int]:
    """
    Быстрое чтение границ листа (read_only) для навигации «перейти к ячейке» до полной отрисовки превью.
    """
    wb = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        if sheet_index < 0 or sheet_index >= len(wb.worksheets):
            raise ValueError(f"Лист с индексом {sheet_index} не найден (всего {len(wb.worksheets)}).")
        ws = wb.worksheets[sheet_index]
        titles = [s.title for s in wb.worksheets]
        smr = max(1, ws.max_row or 1)
        smc = max(1, ws.max_column or 1)
        return titles, smr, smc
    finally:
        wb.close()


def resolve_goto_viewport(
    row: int,
    col: int,
    *,
    viewport_rows: int,
    viewport_cols: int,
    source_max_row: int,
    source_max_column: int,
) -> tuple[int, int]:
    """Центрирует окно просмотра вокруг ячейки (row, col), с отсечением по границам листа."""
    max_start_row = max(1, source_max_row - viewport_rows + 1)
    max_start_col = max(1, source_max_column - viewport_cols + 1)
    top = max(1, min(row - viewport_rows // 2, max_start_row))
    left = max(1, min(col - viewport_cols // 2, max_start_col))
    return top, left


def build_sheet_preview(
    path: Path,
    sheet_index: int,
    *,
    view_top_row: int = 1,
    view_left_col: int = 1,
    viewport_rows: int = DEFAULT_VIEWPORT_ROWS,
    viewport_cols: int = DEFAULT_VIEWPORT_COLS,
) -> dict[str, Any]:
    """
    Фрагмент листа в «окне» [view_top_row, view_left_col] размером viewport_*.
    Учитывает объединённые ячейки в пределах окна.
    """
    if view_top_row < 1 or view_left_col < 1:
        raise ValueError("Координаты окна должны начинаться с 1")
    if viewport_rows < 1 or viewport_cols < 1:
        raise ValueError("Размер окна должен быть ≥ 1")

    wb = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    try:
        titles = [ws.title for ws in wb.worksheets]
        if sheet_index < 0 or sheet_index >= len(titles):
            raise ValueError(f"Лист с индексом {sheet_index} не найдено (всего {len(titles)}).")
        ws = wb.worksheets[sheet_index]

        smr = max(1, ws.max_row or 1)
        smc = max(1, ws.max_column or 1)

        vtr = min(view_top_row, smr)
        vtc = min(view_left_col, smc)

        eff_bottom = min(vtr + viewport_rows - 1, smr)
        eff_right = min(vtc + viewport_cols - 1, smc)

        spans, skip = _apply_merges_viewport(ws, vtr, eff_bottom, vtc, eff_right)

        grid: list[list[PreviewCell]] = []
        for r in range(vtr, eff_bottom + 1):
            row_cells: list[PreviewCell] = []
            for c in range(vtc, eff_right + 1):
                col_w = _column_width_ch(ws, c)
                if (r, c) in skip:
                    row_cells.append(
                        PreviewCell(text="", sheet_row=r, sheet_col=c, skip=True, col_width=col_w),
                    )
                    continue
                cell = ws.cell(row=r, column=c)
                span = spans.get((r, c))
                if span is not None:
                    rs, cs, vr, vc = span
                    text = _cell_text(ws.cell(row=vr, column=vc))
                else:
                    rs, cs = 1, 1
                    text = _cell_text(cell)
                row_cells.append(
                    PreviewCell(
                        text=text,
                        sheet_row=r,
                        sheet_col=c,
                        rowspan=rs,
                        colspan=cs,
                        skip=False,
                        col_width=col_w,
                    ),
                )
            grid.append(row_cells)

        col_letters = [get_column_letter(c) for c in range(vtc, eff_right + 1)]

        truncated_below = eff_bottom < smr
        truncated_right = eff_right < smc
        truncated_above = vtr > 1
        truncated_left = vtc > 1

        return {
            "sheet_titles": titles,
            "sheet_index": sheet_index,
            "sheet_title": titles[sheet_index],
            "grid": grid,
            "col_letters": col_letters,
            "view_top_row": vtr,
            "view_left_col": vtc,
            "view_bottom_row": eff_bottom,
            "view_right_col": eff_right,
            "viewport_rows": viewport_rows,
            "viewport_cols": viewport_cols,
            "truncated_above": truncated_above,
            "truncated_below": truncated_below,
            "truncated_left": truncated_left,
            "truncated_right": truncated_right,
            "rendered_rows": eff_bottom - vtr + 1,
            "rendered_cols": eff_right - vtc + 1,
            "source_max_row": smr,
            "source_max_column": smc,
        }
    finally:
        wb.close()
